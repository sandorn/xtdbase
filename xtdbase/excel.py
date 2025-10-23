#!/usr/bin/env python3
"""
==============================================================
Description  : Excel统一操作类 - 整合openpyxl和pandas的Excel文件读写功能
Develop      : VSCode
Author       : sandorn sandorn@live.cn
LastEditTime : 2025-10-23
Github       : https://github.com/sandorn/xtdbase

本模块提供以下核心功能:
- Excel类:统一封装openpyxl和pandas的Excel操作功能
- 支持精细的单元格操作(openpyxl)和批量数据处理(pandas)
- 支持工作表的创建、删除和切换
- 支持单元格、行、列和整个工作表数据的读写
- 支持数据以字典形式返回,方便处理结构化数据
- 支持多工作表数据处理和文件合并

主要特性:
- 支持上下文管理器,自动处理资源关闭
- 提供灵活的API,支持多种数据读写方式
- 实例方法优先复用当前状态,显式参数可覆盖
- 包含错误处理和类型检查,提高代码健壮性
- 集成日志系统,便于调试和问题追踪
==============================================================
"""

from __future__ import annotations

import os
from collections.abc import Iterator
from pathlib import Path
from typing import IO, Any, cast

import numpy as np
import pandas
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from xtlog import mylog as logger

# 类型别名定义
type PathOrBuffer = str | Path | IO[bytes]


# ============= 数据结构定义 =============


class ColumnMapping(BaseModel):
    """列名映射 - 定义Excel列名与别名的对应关系"""

    column_name: str = Field(description='原始列名')
    column_alias: str = Field(description='显示列名')


class SheetMapping(BaseModel):
    """工作表映射 - 定义文件与工作表名称的对应关系"""

    file_name: str = Field(description='文件名')
    sheet_name: str = Field(description='工作表名称')


class DataCollect(BaseModel):
    """多工作表数据集合 - 封装单个工作表的完整数据和配置"""

    data_list: list[dict] = Field(description='数据列表')
    col_mappings: list[ColumnMapping] = Field(description='列名映射列表')
    sheet_name: str = Field(description='工作表名称')


# ============= 工具函数 =============


def rename_file(file_path: str) -> str:
    """为文件重命名,添加'-excel'后缀

    Args:
        file_path: 原始文件路径

    Returns:
        str: 添加后缀后的文件路径
    """
    file_dir, file_name = os.path.split(file_path)
    base_name, extension = os.path.splitext(file_name)
    return os.path.join(file_dir, f'{base_name}-excel{extension}')


# ============= Excel统一操作类 =============


class Excel:
    """Excel统一操作类 - 整合openpyxl和pandas的功能

    该类提供两种使用模式:
    1. 实例模式 (openpyxl) - 精细的单元格操作、流式读取
    2. 批量模式 (pandas) - 大数据量的批量读写、数据分析

    实例方法优先使用当前实例状态(file, sheet),
    显式传递file或sheet_name参数时可覆盖默认行为。

    Args:
        file: Excel文件路径
        sheet_name: 工作表名称,默认为None(使用活动工作表)
        auto_save_on_exit: 退出上下文管理器时是否自动保存修改,默认为True

    Attributes:
        file: Excel文件路径
        wb: Workbook对象
        sh: 当前工作表对象
        sh_name_list: 所有工作表名称列表
        headers: 表头数据(通过read_header方法设置)
        auto_save_on_exit: 是否在退出时自动保存

    Example:
        >>> # 模式1: 精细操作 (openpyxl)
        >>> with Excel('data.xlsx', 'Sheet1') as excel:
        ...     excel.write_cell(1, 1, '标题')
        ...     data = excel.read_all()
        >>>
        >>> # 模式2: 批量操作 (pandas)
        >>> with Excel('data.xlsx') as excel:
        ...     data = [{'name': 'Alice', 'age': 25}]
        ...     mappings = [ColumnMapping(column_name='name', column_alias='姓名')]
        ...     excel.batch_write(data, mappings)
        >>>
        >>> # 模式3: 文件合并
        >>> Excel.merge_files(['file1.xlsx', 'file2.xlsx'], 'merged.xlsx')
    """

    DEFAULT_SHEET_NAME = 'Sheet1'

    def __init__(self, file: str, sheet_name: str | None = None, auto_save_on_exit: bool = True):
        """初始化Excel实例

        Args:
            file: Excel文件路径
            sheet_name: 工作表名称,默认为None(使用活动工作表)
            auto_save_on_exit: 退出上下文管理器时是否自动保存修改,默认为True

        Raises:
            OSError: 文件创建或加载失败
            PermissionError: 文件权限不足
            TypeError: 工作表类型错误
        """
        self.file = file
        self.auto_save_on_exit = auto_save_on_exit
        logger.info(f'初始化Excel,文件路径: {self.file}')

        # 如果文件不存在,创建新文件
        if not os.path.exists(self.file):
            logger.info(f'文件不存在,创建新文件: {self.file}')

            # 确保目录存在
            file_dir = os.path.dirname(self.file)
            if file_dir and not os.path.exists(file_dir):
                try:
                    os.makedirs(file_dir, exist_ok=True)
                    logger.info(f'创建目录: {file_dir}')
                except OSError as e:
                    logger.error(f'创建目录失败: {e}')
                    raise

            # 创建新文件
            try:
                wb = Workbook()
                wb.save(self.file)
                wb.close()
            except (OSError, PermissionError) as e:
                logger.error(f'创建文件失败: {e}')
                raise

        # 加载工作簿
        try:
            self.wb = load_workbook(self.file)
        except Exception as e:
            logger.error(f'加载工作簿失败: {e}')
            raise

        self.sh_name_list = self.wb.sheetnames

        # 设置当前工作表
        self.sh: Worksheet | None = None
        if sheet_name is None:
            self.sh = self.wb.active
        elif sheet_name not in self.sh_name_list:
            logger.info(f'工作表不存在,创建新工作表: {sheet_name}')
            self.sh = self.wb.create_sheet(sheet_name)
        else:
            self.sh = self.wb[sheet_name]
            logger.info(f'切换到工作表: {sheet_name}')

        if not isinstance(self.sh, Worksheet):
            logger.error('工作表类型错误')
            raise TypeError('工作表类型错误')

        # 初始化表头属性
        self.headers: tuple[Any, ...] | None = None
        # 初始化修改标志
        self._modified = False

    def __enter__(self) -> Excel:
        """支持上下文管理器协议

        Returns:
            Excel: 当前实例
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> bool:
        """支持上下文管理器协议,确保资源正确关闭

        Args:
            exc_type: 异常类型
            exc_val: 异常值
            exc_tb: 异常追踪

        Returns:
            bool: 如果没有异常发生,返回True
        """
        if self._modified and self.auto_save_on_exit:
            self.save_workbook()
            self._modified = False
        self.wb.close()
        logger.info(f'关闭Excel文件: {self.file}')
        return not exc_type

    def _ensure_worksheet(self) -> Worksheet:
        """确保当前工作表有效

        Returns:
            Worksheet: 当前工作表对象

        Raises:
            RuntimeError: 当前没有活动工作表时抛出
        """
        if self.sh is None:
            logger.warning('当前没有活动工作表,返回第一个sheet')
            self.sh = self.wb[self.sh_name_list[0]]
        return self.sh

    # ============= 工作表管理方法 =============

    def create_sheet(self, index: int | None = None, title: str | None = None) -> None:
        """创建新工作表

        Args:
            index: 工作表位置索引
            title: 工作表名称
        """
        self.wb.create_sheet(index=index, title=title)
        self.save_workbook()
        self.sh_name_list = self.wb.sheetnames
        logger.info(f'创建新工作表: 索引={index},标题={title}')

    def remove_sheet(self, sheet_name: str) -> None:
        """删除指定工作表

        Args:
            sheet_name: 要删除的工作表名称
        """
        if sheet_name in self.sh_name_list:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)
            self.save_workbook()
            self.sh_name_list = self.wb.sheetnames
            logger.info(f'删除工作表: {sheet_name}')
        else:
            logger.warning(f'工作表不存在: {sheet_name}')

    def _switch_sheet(self, sheet_name: str, create_if_missing: bool = False) -> None:
        """切换到指定工作表

        Args:
            sheet_name: 工作表名称
            create_if_missing: 如果工作表不存在是否创建，默认为False

        Raises:
            ValueError: 工作表不存在且create_if_missing=False时
            TypeError: 工作表类型错误时
        """
        if sheet_name not in self.sh_name_list:
            if create_if_missing:
                logger.info(f'工作表不存在,创建新工作表: {sheet_name}')
                self.sh = self.wb.create_sheet(sheet_name)
                self.sh_name_list = self.wb.sheetnames
                return
            error_msg = f'工作表不存在: {sheet_name}, 可用工作表: {self.sh_name_list}'
            logger.error(error_msg)
            raise ValueError(error_msg)

        now_sh = self.wb[sheet_name]
        if not isinstance(now_sh, Worksheet):
            error_msg = f'工作表类型错误: {sheet_name},类型: {type(now_sh)}'
            logger.error(error_msg)
            raise TypeError(error_msg)

        self.sh = now_sh
        logger.info(f'切换到工作表: {sheet_name}')

    # ============= 单元格读取方法 (openpyxl) =============

    def read_header(self, sheet_name: str | None = None) -> tuple[Any, ...]:
        """获取表头信息

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            Tuple: 表头数据元组
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()

        # 读取第一行作为表头
        header_row = tuple(sh.iter_rows(max_row=1, values_only=True))
        if header_row:
            self.headers = header_row[0]
            logger.info(f'读取表头成功,包含{len(self.headers)}列')
            return self.headers

        logger.warning('未读取到表头数据')
        return tuple()

    def read_all(self, sheet_name: str | None = None) -> list[list[Any]]:
        """读取工作表中所有数据

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[List[Any]]: 二维列表形式的表格数据
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()

        # 使用 iter_rows + values_only 更高效
        data = [list(row) for row in sh.iter_rows(values_only=True)]
        logger.info(f'读取所有数据成功,包含{len(data)}行')
        return data

    def read_row(self, row: int | str, sheet_name: str | None = None) -> list[Any]:
        """读取指定行数据

        Args:
            row: 行索引(整数)或行标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Any]: 行数据列表
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()
        row_data = [cell.value for cell in sh[row]]
        logger.info(f'读取行{row}数据成功,包含{len(row_data)}列')
        return row_data

    def read_row_dict(self, row: int | str, sheet_name: str | None = None) -> dict[str, Any]:
        """以字典形式读取指定行数据(表头为键)

        Args:
            row: 行索引(整数)或行标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            Dict[str,Any]: 字典形式的行数据
        """
        row_data = self.read_row(row, sheet_name)
        titles = self.read_header(sheet_name)
        result = dict(zip(titles, row_data, strict=False))
        logger.info(f'读取行{row}字典数据成功,包含{len(result)}个字段')
        return result

    def read_col(self, col: int | str, sheet_name: str | None = None) -> list[Any]:
        """读取指定列数据

        Args:
            col: 列索引(整数)或列标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Any]: 列数据列表
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()

        # 处理列标识
        col_index = col if isinstance(col, int) else self.get_column_index(col)

        # 直接从生成器构建列表，避免中间tuple
        col_iter = sh.iter_cols(min_col=col_index, max_col=col_index, values_only=True)
        col_data = list(next(col_iter, []))
        logger.info(f'读取列{col}数据成功,包含{len(col_data)}行')
        return col_data

    def read_all_dict(self, sheet_name: str | None = None) -> list[dict[str, Any]]:
        """以字典列表形式读取所有数据(表头为键)

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Dict[str,Any]]: 字典列表形式的所有数据
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()

        # 使用iter_rows一次性读取，避免重复遍历
        rows_iter = sh.iter_rows(values_only=True)

        # 获取表头
        titles = next(rows_iter, None)
        if not titles:
            logger.warning('未读取到表头数据')
            return []

        # 直接构建字典列表 (显式转换为正确类型)
        data: list[dict[str, Any]] = [cast(dict[str, Any], dict(zip(titles, row, strict=False))) for row in rows_iter]

        logger.info(f'读取所有字典数据成功,包含{len(data)}行记录')
        return data

    def iter_rows_dict(self, sheet_name: str | None = None) -> Iterator[dict[str, Any]]:
        """流式读取字典格式行数据（内存友好，适合大文件）

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Yields:
            dict[str,Any]: 字典格式的行数据

        Example:
            >>> with Excel('large_file.xlsx') as excel:
            ...     for row_dict in excel.iter_rows_dict():
            ...         process_row(row_dict)  # 逐行处理，避免内存溢出
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)

        sh = self._ensure_worksheet()
        rows = sh.iter_rows(values_only=True)

        # 读取表头
        headers = next(rows, None)
        if not headers:
            logger.warning('未读取到表头数据')
            return

        # 逐行返回字典 (显式转换为正确类型)
        for row in rows:
            yield cast(dict[str, Any], dict(zip(headers, row, strict=False)))

    def read_cell(self, row: int, column: int) -> Any:
        """读取指定单元格数据

        Args:
            row: 行索引
            column: 列索引

        Returns:
            Any: 单元格值
        """
        sh = self._ensure_worksheet()
        value = sh.cell(row, column).value
        logger.info(f'读取单元格({row},{column})数据: {value}')
        return value

    # ============= 单元格写入方法 (openpyxl) =============

    def write_cell(self, row: int, column: int, value: Any, auto_save: bool = True) -> None:
        """写入数据到指定单元格

        Args:
            row: 行索引
            column: 列索引
            value: 要写入的值
            auto_save: 是否自动保存，默认为True
        """
        sh = self._ensure_worksheet()
        sh.cell(row, column).value = value
        self._modified = True

        if auto_save:
            self.save_workbook()

        logger.info(f'写入单元格({row},{column})数据: {value}')

    def write_cells(self, cells: list[tuple[int, int, Any]], auto_save: bool = True) -> None:
        """批量写入单元格数据

        Args:
            cells: 单元格数据列表 [(row,col,value),...]
            auto_save: 是否自动保存，默认为True
        """
        sh = self._ensure_worksheet()

        for row, column, value in cells:
            sh.cell(row, column).value = value

        self._modified = True

        if auto_save:
            self.save_workbook()

        logger.info(f'批量写入{len(cells)}个单元格数据')

    def append(self, data_list: list[list[Any]], auto_save: bool = True) -> None:
        """追加写入数据到工作表

        Args:
            data_list: 二维列表形式的数据
            auto_save: 是否自动保存，默认为True
        """
        sh = self._ensure_worksheet()

        for row_data in data_list:
            sh.append(row_data)

        self._modified = True

        if auto_save:
            self.save_workbook()

        logger.info(f'追加写入数据成功,包含{len(data_list)}行')

    # ============= 批量数据操作方法 (pandas) =============

    def batch_write(
        self,
        data: list[dict],
        col_mappings: list[ColumnMapping] | None = None,
        file: str | None = None,
        sheet_name: str | None = None,
        **kwargs,
    ) -> None:
        """批量写入数据到Excel (pandas模式)

        优先使用当前实例状态:
        - 如果file为None,使用self.file
        - 如果sheet_name为None,使用self.sh.title

        Args:
            data: 字典列表数据
            col_mappings: 列名映射,用于重命名列
            file: 目标文件路径,默认为None(使用当前实例文件)
            sheet_name: 工作表名称,默认为None(使用当前工作表)
            **kwargs: 传递给pandas.DataFrame.to_excel的额外参数

        Example:
            >>> # 使用当前实例状态
            >>> with Excel('data.xlsx', 'Sheet1') as excel:
            ...     data = [{'name': 'Alice', 'age': 25}]
            ...     mappings = [ColumnMapping(column_name='name', column_alias='姓名')]
            ...     excel.batch_write(data, mappings)
            >>>
            >>> # 显式指定文件和工作表
            >>> with Excel('temp.xlsx') as excel:
            ...     excel.batch_write(data, mappings, file='output.xlsx', sheet_name='Users')
        """
        target_file = file or self.file
        target_sheet = sheet_name or (self.sh.title if self.sh else self.DEFAULT_SHEET_NAME)

        logger.info(f'批量写入数据: 文件={target_file}, 工作表={target_sheet}, 数据行数={len(data)}')

        try:
            # 如果有列映射,在创建DataFrame时直接使用重命名后的列名
            if col_mappings:
                col_dict = {cm.column_name: cm.column_alias for cm in col_mappings}
                renamed_data = [{col_dict.get(k, k): v for k, v in row.items()} for row in data]
                df = pandas.DataFrame(data=renamed_data)
            else:
                df = pandas.DataFrame(data=data)

            # 判断是否写入当前实例文件
            if target_file == self.file:
                # 写入当前文件,需要使用ExcelWriter追加模式
                with pandas.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:  # type: ignore[arg-type]
                    df.to_excel(writer, sheet_name=target_sheet, index=False, **kwargs)
                # 重新加载工作簿以保持状态同步
                current_sheet_title = self.sh.title if self.sh else None
                self.wb = load_workbook(self.file)
                self.sh_name_list = self.wb.sheetnames
                # 尝试恢复到原工作表,如果不存在则使用active
                if current_sheet_title and current_sheet_title in self.sh_name_list:
                    self.sh = self.wb[current_sheet_title]
                else:
                    self.sh = self.wb.active
            else:
                # 写入不同文件,直接创建新文件
                with pandas.ExcelWriter(target_file) as writer:  # type: ignore[arg-type]
                    df.to_excel(writer, sheet_name=target_sheet, index=False, **kwargs)

            logger.info(f'批量写入数据成功: {target_file}')
        except PermissionError as e:
            logger.error(f'没有权限写入文件, 请检查文件是否被占用: {e!s}')
            raise
        except Exception as e:
            logger.error(f'批量写入数据失败: {e!s}')
            raise

    def batch_read(
        self,
        file: str | None = None,
        sheet_name: str | None = None,
        col_mappings: list[ColumnMapping] | None = None,
        all_col: bool = True,
        header: int = 0,
        nan_replace=None,
        **kwargs,
    ) -> list[dict]:
        """批量读取数据 (pandas模式)

        优先使用当前实例状态:
        - 如果file为None,使用self.file
        - 如果sheet_name为None,使用self.sh.title

        Args:
            file: 源文件路径,默认为None(使用当前实例文件)
            sheet_name: 工作表名称,默认为None(使用当前工作表)
            col_mappings: 列字段映射,用于重命名列
            all_col: True返回所有列信息,False则只返回col_mapping对应的字段信息
            header: 指定表头所在行,默认为0(第一行)
            nan_replace: NaN值的替换值,默认为None
            **kwargs: 传递给pandas.read_excel的额外参数

        Returns:
            list[dict]: 读取的数据列表,每行数据以字典形式表示
        """
        target_file = file or self.file
        target_sheet = sheet_name or (self.sh.title if self.sh else self.DEFAULT_SHEET_NAME)

        logger.info(f'批量读取数据: 文件={target_file}, 工作表={target_sheet}')

        try:
            # 构建列名映射字典
            col_dict = {cm.column_name: cm.column_alias for cm in col_mappings} if col_mappings else None

            # 确定要读取的列
            use_cols = None
            if not all_col:
                use_cols = list(col_dict) if col_dict else None

            # 读取Excel数据
            df = pandas.read_excel(
                target_file,
                sheet_name=target_sheet,
                usecols=use_cols,
                header=header,
                **kwargs,
            )

            # 替换NaN值(仅在需要时执行)
            if nan_replace is not None:
                df.replace(np.nan, nan_replace, inplace=True)

            # 重命名列
            if col_dict:
                df.rename(columns=col_dict, inplace=True)

            # 转换为字典列表
            result = df.to_dict('records')
            logger.info(f'批量读取数据成功,包含{len(result)}条记录')
            return result
        except Exception as e:
            logger.error(f'批量读取数据失败: {e!s}')
            raise

    def multi_sheet_write(self, data_collects: list[DataCollect], file: str | None = None, **kwargs) -> None:
        """将多个数据列表写入同一个Excel文件的不同工作表

        Args:
            data_collects: 数据集列表,每个元素包含一个工作表的数据
            file: 目标文件路径,默认为None(使用当前实例文件)
            **kwargs: 传递给pandas.ExcelWriter的额外参数
        """
        target_file = file or self.file

        logger.info(f'多工作表写入: 文件={target_file}, 工作表数量={len(data_collects)}')

        try:
            with pandas.ExcelWriter(target_file) as writer:  # type: ignore[arg-type]
                for data_collect in data_collects:
                    # 处理列映射
                    if data_collect.col_mappings:
                        col_dict = {cm.column_name: cm.column_alias for cm in data_collect.col_mappings}
                        renamed_data = [{col_dict.get(k, k): v for k, v in row.items()} for row in data_collect.data_list]
                        df = pandas.DataFrame(data=renamed_data)
                    else:
                        df = pandas.DataFrame(data=data_collect.data_list)

                    df.to_excel(writer, sheet_name=data_collect.sheet_name, index=False, **kwargs)
                    logger.info(f'写入工作表[{data_collect.sheet_name}]成功,包含{len(data_collect.data_list)}行')

            logger.info(f'多工作表写入完成: {target_file}')

            # 如果写入当前实例文件,重新加载工作簿
            if target_file == self.file:
                current_sheet_title = self.sh.title if self.sh else None
                self.wb = load_workbook(self.file)
                self.sh_name_list = self.wb.sheetnames
                # 尝试恢复到原工作表,如果不存在则使用第一个工作表
                if current_sheet_title and current_sheet_title in self.sh_name_list:
                    self.sh = self.wb[current_sheet_title]
                elif self.sh_name_list:
                    self.sh = self.wb[self.sh_name_list[0]]
                else:
                    self.sh = self.wb.active
        except PermissionError as e:
            logger.error(f'没有权限写入文件,请检查文件是否被占用: {e!s}')
            raise
        except Exception as e:
            logger.error(f'多工作表写入失败: {e!s}')
            raise

    @staticmethod
    def merge_files(
        input_files: list[str],
        output_file: str,
        sheet_mappings: list[SheetMapping] | None = None,
        validate_files: bool = True,
        **kwargs,
    ) -> None:
        """合并多个Excel文件到一个文件中,每个文件对应一个工作表

        这是一个静态方法,不依赖实例状态,可以直接调用。

        Args:
            input_files: 待合并的Excel文件列表
            output_file: 输出文件路径
            sheet_mappings: 文件工作表映射,默认使用文件名
            validate_files: 是否预先验证文件存在,默认为True
            **kwargs: 传递给pandas.ExcelWriter的额外参数

        Example:
            >>> Excel.merge_files(['file1.xlsx', 'file2.xlsx'], 'merged.xlsx')
        """
        try:
            sheet_mappings = sheet_mappings or []
            sheet_dict = {sheet_mapping.file_name: sheet_mapping.sheet_name for sheet_mapping in sheet_mappings}

            # 预先验证文件列表
            if validate_files:
                valid_files = [f for f in input_files if os.path.exists(f)]
                invalid_count = len(input_files) - len(valid_files)
                if invalid_count > 0:
                    logger.warning(f'跳过 {invalid_count} 个不存在的文件')
            else:
                valid_files = input_files

            logger.info(f'开始合并Excel文件, 有效文件数: {len(valid_files)}, 输出文件: {output_file}')

            with pandas.ExcelWriter(output_file, engine_kwargs=kwargs) as writer:
                for file in valid_files:
                    try:
                        df = pandas.read_excel(file)
                        file_name = os.path.basename(file)
                        sheet_name = sheet_dict.get(file_name, file_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        logger.info(f'成功合并文件[{file}]到工作表[{sheet_name}]')
                    except PermissionError as e:
                        logger.error(f'没有权限读取文件[{file}]: {e!s}')
                        continue
                    except Exception as e:
                        logger.error(f'合并文件[{file}]失败: {e!s}')
                        continue

            logger.info(f'Excel文件合并完成, 输出文件: {output_file}')
        except PermissionError as e:
            logger.error(f'没有权限写入输出文件{output_file}: {e!s}')
            raise
        except Exception as e:
            logger.error(f'Excel文件合并失败: {e!s}')
            raise

    # ============= 工具方法 =============

    def save_workbook(self, excel_path: str | None = None) -> None:
        """保存工作簿

        Args:
            excel_path: 保存路径,默认为None(保存到原文件)
        """
        save_path = excel_path or self.file
        self.wb.save(save_path)
        logger.info(f'保存工作簿成功: {save_path}')

    @staticmethod
    def get_column_index(col_str: str) -> int:
        """将列标识转换为列索引

        Args:
            col_str: 列标识(如'A','B','AA')

        Returns:
            int: 列索引
        """
        return column_index_from_string(col_str)

    @staticmethod
    def get_column_letter(col_num: int) -> str:
        """将列索引转换为列标识

        Args:
            col_num: 列索引

        Returns:
            str: 列标识
        """
        return get_column_letter(col_num)
